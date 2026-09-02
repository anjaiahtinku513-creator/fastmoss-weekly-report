#!/usr/bin/env python3
"""Run a deterministic, synthetic smoke test for prepare_report.py."""

from __future__ import annotations

import csv
import json
import tempfile
from pathlib import Path

from prepare_report import (
    build_product_identifier_index,
    load_rows,
    load_structured_product_catalog,
    main,
    normalize_candidates,
    parse_number,
    percentile_map,
    product_match,
)


ANGLES = ["problem-solution", "fit-proof", "styling", "unboxing", "comparison", "material-detail"]


def build_products(path: Path) -> None:
    fields = [
        "sku", "title", "market", "category", "product_type", "pain_points", "selling_points",
        "proof_actions", "fit", "colors", "neckline", "sleeve", "fabric", "scenarios",
        "price", "stock", "margin", "prohibited_claims", "image_paths",
    ]
    rows = []
    for index in range(1, 7):
        rows.append({
            "sku": f"SKU-{index:03d}",
            "title": f"Soft everyday top {index}",
            "market": "US",
            "category": "Women Tops",
            "product_type": "casual top",
            "pain_points": "clings at the stomach; hard to style",
            "selling_points": "soft drape; forgiving fit; easy outfit",
            "proof_actions": "side turn and fabric pinch",
            "fit": "relaxed",
            "colors": "black; blue",
            "neckline": "crew neck",
            "sleeve": "short sleeve",
            "fabric": "soft knit",
            "scenarios": "work; weekend",
            "price": "29.99",
            "stock": "120",
            "margin": "42%",
            "prohibited_claims": "slimming guarantee",
            "image_paths": "",
        })
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def build_candidates(path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("Smoke test requires bundled openpyxl") from exc
    headers = [
        "视频ID", "视频链接", "达人", "发布时间", "市场", "类目", "商品ID", "商品名称",
        "价格", "播放量", "点赞数", "评论数", "分享数", "销量", "GMV", "播放增长",
        "销量增长", "视频时长", "视频文案", "转录", "采集方式", "采集时间", "开头钩子",
        "购买痛点", "证明动作", "CTA", "内容角度", "证据状态", "市场口播", "中文翻译",
        "封面文案", "标签",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FastMoss Export"
    sheet.append(headers)
    for index in range(12):
        sku = f"SKU-{index % 6 + 1:03d}"
        sheet.append([
            str(7400000000000000000 + index),
            f"https://www.tiktok.com/@creator{index}/video/{7400000000000000000 + index}",
            f"creator{index}",
            f"2026-08-04 {8 + index % 8:02d}:15:00",
            "US",
            "Women Tops",
            sku,
            f"{sku} soft casual top",
            "$29.99",
            f"{100 + index * 25}K",
            str(9000 + index * 500),
            str(300 + index * 20),
            str(150 + index * 15),
            str(80 + index * 8),
            f"${2500 + index * 350}",
            f"{20 + index}%",
            f"{10 + index}%",
            "15",
            "A casual top that does not cling at the stomach and is easy to style.",
            "Watch the side turn and fabric pinch. The soft drape makes an easy outfit.",
            "official_export",
            "2026-08-05 07:00:00",
            "If your tee clings right here, watch this.",
            "clings at the stomach",
            "side turn and fabric pinch",
            "I linked it below.",
            ANGLES[index % len(ANGLES)],
            "available",
            "If your tee clings here, watch the side turn. This drape is why I kept it. Linked below.",
            "如果你的T恤总贴肚子，看这个侧身展示。这个垂坠感就是我留下它的原因，链接在下方。",
            "The side-turn test",
            "#TikTokShop #WomensTops #StyleFinds #Imily Bela #TryOn",
        ])
    workbook.save(path)


def build_creator_profiles(path: Path) -> None:
    profiles = []
    for index in range(11):
        profiles.append({
            "creator_name": f"creator{index}",
            "creator_followers": 50000 if index == 10 else 1000 + index,
            "creator_followers_as_of": "2026-08-05",
            "creator_profile_source": "synthetic_exact_handle_profile",
            "creator_profile_source_url": f"https://example.com/tiktok/creator{index}",
        })
    path.write_text(json.dumps({"profiles": profiles}, ensure_ascii=False, indent=2), encoding="utf-8")


def build_structured_catalog(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    master = workbook.active
    master.title = "\u5546\u54c1\u4e3b\u8868"
    master.append(["market", "sku", "display_name", "title", "category", "product_type", "pain_points", "selling_points", "proof_actions", "fabric", "colors", "normal_price", "promo_price", "margin", "catalog_status"])
    master.append(["US", "SKU-001", "US top", "Soft US top", "Women Tops", "casual top", "stomach cling", "soft drape", "side turn", "cotton", "blue", 29.99, 24.99, 0.42, "ready_for_semantic_match"])
    master.append(["DE", "SKU-001", "DE top", "Weiches DE Top", "Damen Oberteile", "Freizeitshirt", "liegt am Bauch an", "weicher Fall", "Seitendrehung", "Baumwolle", "blau", 31.99, 27.99, 0.35, "ready_for_semantic_match"])
    master.append(["DE", "SKU-002", "DE no stock", "DE no stock", "Damen Oberteile", "Freizeitshirt", "", "weich", "", "Polyester", "schwarz", 21.99, None, 0.30, "usable_with_review"])

    aliases = workbook.create_sheet("\u6807\u8bc6\u522b\u540d\u8868")
    aliases.append(["market", "sku", "old_skus", "tiktok_product_ids", "primary_tiktok_product_id", "amazon_asins", "preferred_asin"])
    aliases.append(["US", "SKU-001", "OLD-US", "US-PID", "US-PID", "B000000001", "B000000001"])
    aliases.append(["DE", "SKU-001", "OLD-DE", "DE-PID", "DE-PID", "B000000002", "B000000002"])
    aliases.append(["DE", "SKU-002", "", "DE-PID-2", "DE-PID-2", "", ""])

    inventory = workbook.create_sheet("\u5e02\u573a\u5e93\u5b58\u8868")
    inventory.append(["market", "sku", "source_sku", "snapshot_date", "stock_status", "available_stock", "in_transit_stock"])
    inventory.append(["US", "SKU-001", "SKU-001", "2026-08-04", "in_stock", 120, 0])
    inventory.append(["DE", "SKU-001", "SKU-001", "2026-08-04", "in_stock", 45, 5])
    inventory.append(["DE", "SKU-002", "SKU-002", "2026-08-04", "out_of_stock", 0, 0])
    workbook.save(path)


def read_ids(path: Path, filename: str) -> list[str]:
    payload = json.loads((path / filename).read_text(encoding="utf-8"))
    return [row["video_id"] for row in payload]


def run() -> None:
    assert parse_number("1.2K") == 1200
    assert parse_number("3.4万") == 34000
    assert abs((parse_number("42%") or 0) - 0.42) < 1e-9
    assert abs((parse_number("15,21 \u20ac") or 0) - 15.21) < 1e-9
    official = normalize_candidates([{
        "商品标题": "Official export tee",
        "商品售价": "$19.99",
        "商品分类": "Women Tops",
        "视频标题": "Official caption",
        "视频播放量": 1234,
        "视频销量": 12,
        "视频地址": "https://www.tiktok.com/@official_creator/video/7668360702578429198",
        "视频总销量": 20,
        "视频总销售额": 399.8,
        "总播放量": 2222,
        "总点赞量": 100,
        "TikTok商品链接": "https://shop.tiktok.com/view/product/1731884684855317262?region=US",
        "FastMoss商品详情页链接": "https://www.fastmoss.com/en/e-commerce/detail/1731884684855317262",
    }], "US", "official_export")
    assert len(official) == 1
    assert official[0]["creator_name"] == "official_creator"
    assert official[0]["product_id"] == "1731884684855317262"
    assert official[0]["source_tier"] == "official_export"
    assert official[0]["engagement_rate"] is None
    assert official[0]["product_total_gmv"] == 399.8
    tied = percentile_map([
        {"_key": "a", "sales": 1.0},
        {"_key": "b", "sales": 1.0},
    ], "sales")
    assert tied == {"a": 50.0, "b": 50.0}
    with tempfile.TemporaryDirectory(prefix="fastmoss-skill-smoke-") as temp:
        root = Path(temp)
        candidates = root / "candidates.xlsx"
        products = root / "products.csv"
        profiles = root / "creator-profiles.json"
        json_fixture = root / "fixture.json"
        jsonl_fixture = root / "fixture.jsonl"
        bad_products = root / "bad-products.csv"
        structured_catalog = root / "structured-catalog.xlsx"
        out_a, out_b, out_fail = root / "run-a", root / "run-b", root / "run-fail"
        build_candidates(candidates)
        build_products(products)
        build_creator_profiles(profiles)
        build_structured_catalog(structured_catalog)
        bad_products.write_text("title,market\nNo concrete SKU,US\n", encoding="utf-8")
        json_fixture.write_text(json.dumps({"data": [{"video_id": "json-1"}]}), encoding="utf-8")
        jsonl_fixture.write_text(json.dumps({"video_id": "jsonl-1"}) + "\n", encoding="utf-8")
        assert load_rows(json_fixture)[0]["video_id"] == "json-1"
        assert load_rows(jsonl_fixture)[0]["video_id"] == "jsonl-1"
        us_products, us_meta = load_structured_product_catalog(structured_catalog, "US")
        de_products, de_meta = load_structured_product_catalog(structured_catalog, "DE")
        assert len(us_products) == 1 and us_products[0]["market"] == "US" and us_products[0]["stock"] == 120
        assert len(de_products) == 1 and de_products[0]["market"] == "DE" and de_products[0]["stock"] == 45
        assert us_meta["out_of_stock_filtered"] == 0
        assert de_meta["out_of_stock_filtered"] == 1
        de_all, _ = load_structured_product_catalog(structured_catalog, "DE", allow_out_of_stock=True)
        assert {product["sku"] for product in de_all} == {"SKU-001", "SKU-002"}
        exact_product, exact_score, reasons = product_match({
            "market": "US", "product_id": "US-PID", "product_url": "", "product_title": "",
            "category": "", "caption": "", "transcript": "", "pain_point": "", "hook": "",
            "proof_action": "", "price": None,
        }, us_products)
        assert exact_product and exact_product["sku"] == "SKU-001" and exact_score >= 95
        assert any("product ID" in reason for reason in reasons)
        identifier_index = build_product_identifier_index(us_products)
        no_exact_product, no_exact_score, no_exact_reasons = product_match({
            "market": "US", "product_id": "UNKNOWN-PID", "product_url": "", "product_title": "Soft US top",
            "category": "Women Tops", "caption": "", "transcript": "", "pain_point": "", "hook": "",
            "proof_action": "", "price": None,
        }, us_products, identifier_index, True)
        assert no_exact_product is None and no_exact_score == 0
        assert "no exact" in no_exact_reasons[0]
        ambiguous_product, ambiguous_score, ambiguous_reasons = product_match({
            "market": "US", "product_id": "US-PID", "product_url": "", "product_title": "Soft US top",
            "category": "Women Tops", "caption": "", "transcript": "", "pain_point": "", "hook": "",
            "proof_action": "", "price": None, "catalog_match_status": "ambiguous_exact_id",
        }, us_products, identifier_index, True)
        assert ambiguous_product is None and ambiguous_score == 0
        assert "ambiguous" in ambiguous_reasons[0]
        duplicate_identifier_product = dict(us_products[0])
        duplicate_identifier_product["sku"] = "SKU-999"
        duplicate_index = build_product_identifier_index([us_products[0], duplicate_identifier_product])
        duplicate_product, duplicate_score, duplicate_reasons = product_match({
            "market": "US", "product_id": "US-PID", "product_url": "", "product_title": "",
            "category": "", "caption": "", "transcript": "", "pain_point": "", "hook": "",
            "proof_action": "", "price": None,
        }, [us_products[0], duplicate_identifier_product], duplicate_index, True)
        assert duplicate_product is None and duplicate_score == 0
        assert "multiple" in duplicate_reasons[0]
        common = [
            "--candidates", str(candidates), "--products", str(products), "--market", "US",
            "--analysis-date", "2026-08-04", "--top", "10", "--min-candidates", "12",
            "--min-match", "20", "--creator-profiles", str(profiles), "--max-creator-followers", "30000",
        ]
        assert main([*common, "--output-dir", str(out_a)]) == 0
        assert main([*common, "--output-dir", str(out_b)]) == 0
        manifest = json.loads((out_a / "run-manifest.json").read_text(encoding="utf-8"))
        assert manifest["passed"], manifest
        assert manifest["selected_count"] == 10
        assert manifest["gates"]["selected_creator_size_verified"]
        assert manifest["creator_size_statuses"]["over_limit"] == 1
        assert manifest["creator_size_statuses"]["unverified"] == 1
        assert manifest["content_angle_count"] >= 4
        top10 = json.loads((out_a / "top10.json").read_text(encoding="utf-8"))
        assert all(row["creator_followers"] <= 30000 for row in top10)
        assert read_ids(out_a, "ranked-candidates.json") == read_ids(out_b, "ranked-candidates.json")
        assert read_ids(out_a, "top10.json") == read_ids(out_b, "top10.json")
        assert "#Imily Bela" in (out_a / "daily-report.md").read_text(encoding="utf-8")
        assert (out_a / "failure-log.jsonl").read_text(encoding="utf-8") == ""
        assert main([
            "--candidates", str(candidates), "--products", str(products), "--market", "US",
            "--analysis-date", "2026-08-04", "--output-dir", str(out_fail), "--top", "10",
            "--min-candidates", "100", "--min-match", "20", "--creator-profiles", str(profiles), "--max-creator-followers", "30000",
        ]) == 0
        failed_manifest = json.loads((out_fail / "run-manifest.json").read_text(encoding="utf-8"))
        assert not failed_manifest["passed"]
        assert "candidate_floor_not_met" in (out_fail / "failure-log.jsonl").read_text(encoding="utf-8")
        try:
            main([
                "--candidates", str(candidates), "--products", str(bad_products), "--market", "US",
                "--analysis-date", "2026-08-04", "--output-dir", str(root / "bad-run"),
            ])
        except ValueError as exc:
            assert "concrete SKUs" in str(exc)
        else:
            raise AssertionError("Product input without a concrete SKU was accepted")
        print(json.dumps({
            "passed": True,
            "candidate_count": manifest["candidate_count"],
            "selected_count": manifest["selected_count"],
            "content_angle_count": manifest["content_angle_count"],
            "deterministic": True,
            "xlsx_parse": True,
            "json_parse": True,
            "jsonl_parse": True,
            "failure_logging": True,
            "incomplete_product_rejected": True,
            "structured_catalog_market_isolation": True,
            "structured_catalog_stock_gate": True,
            "structured_catalog_identifier_match": True,
            "creator_size_gate": True,
        }, ensure_ascii=False))


if __name__ == "__main__":
    run()
