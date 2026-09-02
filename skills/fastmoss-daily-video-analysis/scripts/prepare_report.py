#!/usr/bin/env python3
"""Normalize FastMoss data, rank candidates, match SKUs, and build a report."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone, tzinfo
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError


MISSING = {"", "-", "--", "n/a", "na", "none", "null", "unknown", "未知", "未公开", "隐藏"}
MARKET_TIMEZONES = {"US": "America/New_York", "DE": "Europe/Berlin"}
FALLBACK_MARKET_OFFSETS = {
    "US": {True: -4, False: -5},
    "DE": {True: 2, False: 1},
}

CANDIDATE_ALIASES = {
    "store_id": ["store_id", "store", "shop", "shop_id", "店铺", "店铺id", "店铺名称"],
    "video_id": ["video_id", "videoid", "视频id", "作品id", "tiktokvideoid"],
    "video_url": ["video_url", "videourl", "url", "link", "视频链接", "作品链接", "tiktok链接"],
    "creator_name": ["creator_name", "creator", "author", "达人", "达人昵称", "创作者", "账号"],
    "creator_id": ["creator_id", "creatorid", "达人id", "作者id"],
    "creator_followers": ["creator_followers", "follower_count", "followers", "达人粉丝数", "粉丝数"],
    "creator_followers_as_of": ["creator_followers_as_of", "followers_as_of", "follower_updated_at", "粉丝数日期", "粉丝更新时间"],
    "creator_profile_source": ["creator_profile_source", "follower_source", "粉丝证据来源"],
    "creator_profile_source_url": ["creator_profile_source_url", "creator_profile_url", "profile_url", "follower_source_url", "达人主页", "账号链接"],
    "published_at": ["published_at", "publish_time", "published", "发布时间", "发布日期", "创建时间"],
    "market": ["market", "region", "country", "国家", "地区", "市场"],
    "category": ["category", "类目", "商品类目", "视频类目"],
    "product_id": ["product_id", "productid", "商品id", "产品id"],
    "product_title": ["product_title", "product", "productname", "商品", "商品名称", "产品名称"],
    "product_url": ["product_url", "productlink", "商品链接", "产品链接"],
    "price": ["price", "商品价格", "价格", "售价"],
    "currency": ["currency", "币种", "货币"],
    "views": ["views", "view_count", "播放量", "观看量", "播放次数"],
    "likes": ["likes", "like_count", "点赞", "点赞数"],
    "comments": ["comments", "comment_count", "评论", "评论数"],
    "shares": ["shares", "share_count", "分享", "分享数"],
    "sales": ["sales", "sold", "orders", "销量", "销售量", "订单量", "带货销量"],
    "gmv": ["gmv", "revenue", "销售额", "成交额", "带货销售额"],
    "engagement_rate": ["engagement_rate", "engagement", "互动率"],
    "view_growth": ["view_growth", "views_growth", "播放增量", "播放增长", "播放增长率"],
    "sales_growth": ["sales_growth", "销量增量", "销量增长", "销量增长率"],
    "duration_sec": ["duration_sec", "duration", "视频时长", "时长"],
    "caption": ["caption", "description", "title", "文案", "视频文案", "标题"],
    "transcript": ["transcript", "voiceover", "script", "转录", "口播", "脚本"],
    "thumbnail_url": ["thumbnail_url", "cover", "cover_url", "封面", "封面链接"],
    "source_tier": ["source_tier", "source", "数据来源", "采集方式"],
    "captured_at": ["captured_at", "capture_time", "采集时间", "抓取时间"],
    "hook": ["hook", "开头钩子", "钩子"],
    "pain_point": ["pain_point", "pain", "购买痛点", "痛点"],
    "proof_action": ["proof_action", "proof", "证明动作", "验证动作"],
    "cta": ["cta", "行动号召", "购买引导"],
    "content_angle": ["content_angle", "angle", "内容角度", "视频角度"],
    "source_video_structure": ["source_video_structure", "actual_video_structure", "video_structure", "source_structure", "原视频结构", "爆款视频结构", "实际视频结构"],
    "source_voiceover_full": ["source_voiceover_full", "full_voiceover", "source_transcript", "完整口播", "原视频口播", "原视频完整口播"],
    "source_voiceover_zh": ["source_voiceover_zh", "full_voiceover_zh", "source_voiceover_translation_zh", "原视频口播中文", "完整口播中文", "中文翻译"],
    "voiceover_source": ["voiceover_source", "transcript_source", "口播来源"],
    "risk_note": ["risk_note", "risk", "风险", "风险提示"],
    "evidence_status": ["evidence_status", "证据状态"],
    "voiceover": ["voiceover", "market_voiceover", "市场口播"],
    "voiceover_zh": ["voiceover_zh", "口播中文"],
    "cover_copy": ["cover_copy", "封面文案"],
    "hashtags": ["hashtags", "标签"],
}

CREATOR_PROFILE_ALIASES = {
    "creator_name": ["creator_name", "creator", "username", "handle", "author", "达人", "创作者", "账号"],
    "creator_followers": ["creator_followers", "follower_count", "followers", "达人粉丝数", "粉丝数"],
    "creator_followers_as_of": ["creator_followers_as_of", "followers_as_of", "updated_at", "as_of", "captured_at", "粉丝数日期", "粉丝更新时间"],
    "creator_profile_source": ["creator_profile_source", "source", "source_name", "evidence_source", "证据来源"],
    "creator_profile_source_url": ["creator_profile_source_url", "profile_url", "source_url", "url", "达人主页", "账号链接"],
}

# Current FastMoss official-export headers. Aggregate product metrics stay
# separate from video-level metrics so they cannot inflate video scoring.
CANDIDATE_ALIASES["product_title"].append("商品标题")
CANDIDATE_ALIASES["price"].append("商品售价")
CANDIDATE_ALIASES["category"].append("商品分类")
CANDIDATE_ALIASES["caption"].append("视频标题")
CANDIDATE_ALIASES["views"].append("视频播放量")
CANDIDATE_ALIASES["sales"].append("视频销量")
CANDIDATE_ALIASES["video_url"].append("视频地址")
CANDIDATE_ALIASES["product_url"].append("TikTok商品链接")
CANDIDATE_ALIASES.update({
    "product_image_url": ["product_image_url", "商品封面"],
    "fastmoss_product_url": ["fastmoss_product_url", "FastMoss商品详情页链接"],
    "product_total_sales": ["product_total_sales", "视频总销量"],
    "product_total_gmv": ["product_total_gmv", "视频总销售额"],
    "product_total_views": ["product_total_views", "总播放量"],
    "product_total_likes": ["product_total_likes", "总点赞量"],
    "product_estimated_listed_at": ["product_estimated_listed_at", "预估商品上架时间"],
    "catalog_match_status": ["catalog_match_status"],
})

PRODUCT_ALIASES = {
    "store_id": ["store_id", "store", "shop", "shop_id", "店铺", "店铺id", "店铺名称", "店铺名"],
    "sku": ["sku", "seller_sku", "商品sku", "产品sku", "货号", "款号"],
    "title": ["title", "product_title", "name", "商品名称", "产品名称", "产品标题", "品名"],
    "market": ["market", "region", "country", "市场", "国家", "地区"],
    "category": ["category", "类目", "商品类目", "分类id"],
    "product_type": ["product_type", "type", "产品类型", "商品类型", "款式"],
    "pain_points": ["pain_points", "pain", "买家痛点", "痛点", "购买顾虑"],
    "selling_points": ["selling_points", "features", "卖点", "核心卖点"],
    "proof_actions": ["proof_actions", "proof", "证明动作", "展示动作"],
    "fit": ["fit", "版型", "合身度"],
    "colors": ["colors", "color", "颜色", "色系"],
    "neckline": ["neckline", "领型", "领口"],
    "sleeve": ["sleeve", "袖型", "袖长"],
    "fabric": ["fabric", "material", "面料", "材质"],
    "scenarios": ["scenarios", "occasion", "场景", "适用场景"],
    "price": ["price", "售价", "价格"],
    "stock": ["stock", "inventory", "库存", "可售库存"],
    "margin": ["margin", "profit_margin", "毛利", "毛利率"],
    "prohibited_claims": ["prohibited_claims", "forbidden_claims", "禁用话术", "禁止宣称"],
    "image_paths": ["image_paths", "images", "图片", "产品图", "素材路径", "主图(url)地址", "主图", "附图一"],
}

PRODUCT_ALIASES.update({
    "display_name": ["display_name"],
    "title_status": ["title_status"],
    "catalog_segment": ["catalog_segment"],
    "planning_type": ["planning_type"],
    "style_label": ["style_label"],
    "season": ["season"],
    "operations_note": ["operations_note"],
    "old_skus": ["old_skus", "old_sku", "aliases"],
    "product_ids": ["product_ids", "tiktok_product_ids", "产品id", "全球产品id"],
    "primary_product_id": ["primary_product_id", "primary_tiktok_product_id"],
    "asins": ["asins", "amazon_asins"],
    "preferred_asin": ["preferred_asin"],
    "normal_price": ["normal_price"],
    "promo_price": ["promo_price"],
    "currency": ["currency"],
    "weight_g": ["weight_g"],
    "cost_cny": ["cost_cny"],
    "profit_cny": ["profit_cny"],
    "gmv": ["gmv"],
    "units_sold": ["units_sold"],
    "video_gmv": ["video_gmv"],
    "product_card_gmv": ["product_card_gmv"],
    "catalog_status": ["catalog_status"],
    "missing_fields": ["missing_fields"],
})

STRUCTURED_CATALOG_SHEETS = {
    "product_master": {"\u5546\u54c1\u4e3b\u8868", "product_master", "Product Master"},
    "identifier_aliases": {"\u6807\u8bc6\u522b\u540d\u8868", "identifier_aliases", "Identifier Aliases"},
    "market_inventory": {"\u5e02\u573a\u5e93\u5b58\u8868", "market_inventory", "Market Inventory"},
}

NUMERIC_CANDIDATE_FIELDS = {
    "price", "views", "likes", "comments", "shares", "sales", "gmv", "creator_followers",
    "engagement_rate", "view_growth", "sales_growth", "duration_sec",
    "product_total_sales", "product_total_gmv", "product_total_views", "product_total_likes",
}
CORE_FIELDS = ["identity", "creator_name", "published_at", "market", "product_title", "views", "commerce"]
CREATOR_OUTPUT_FIELDS = [
    "creator_size_status", "creator_size_rejection_reason", "creator_profile_match_basis",
]

ANGLE_KEYWORDS = {
    "problem-solution": ["problem", "struggle", "hate", "finally", "解决", "显肚", "遮", "痛点", "problemzone"],
    "fit-proof": ["fit", "try on", "try-on", "body", "size", "stretch", "合身", "试穿", "身材", "版型"],
    "styling": ["style", "outfit", "ways to wear", "look", "搭配", "穿搭", "造型"],
    "unboxing": ["unbox", "package", "haul", "first impression", "开箱", "拆包", "到货"],
    "comparison": ["before", "after", "versus", "vs", "compare", "对比", "前后"],
    "material-detail": ["fabric", "soft", "texture", "seam", "detail", "面料", "柔软", "细节", "走线"],
    "value": ["price", "deal", "worth", "sale", "budget", "价格", "折扣", "值不值"],
    "social-proof": ["review", "viral", "everyone", "comment", "评价", "爆款", "评论"],
}


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def alias_lookup(aliases: dict[str, list[str]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for canonical, names in aliases.items():
        for name in [canonical, *names]:
            result[normalize_key(name)] = canonical
    return result


def canonicalize_row(row: dict[str, Any], aliases: dict[str, list[str]]) -> dict[str, Any]:
    lookup = alias_lookup(aliases)
    result: dict[str, Any] = {}
    for key, value in row.items():
        canonical = lookup.get(normalize_key(key))
        if canonical and (canonical not in result or is_missing(result[canonical])):
            result[canonical] = value
    return result


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip().lower() in MISSING


def clean_text(value: Any) -> str:
    if is_missing(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _parse_number_legacy(value: Any) -> float | None:
    if is_missing(value):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = unicodedata.normalize("NFKC", str(value)).strip().lower().replace(",", "")
    is_percent = "%" in text
    multipliers = {"k": 1_000, "m": 1_000_000, "b": 1_000_000_000, "万": 10_000, "亿": 100_000_000}
    multiplier = 1.0
    for suffix, factor in multipliers.items():
        if suffix in text:
            multiplier = factor
            break
    match = re.search(r"[-+]?\d*\.?\d+", text)
    if not match:
        return None
    number = float(match.group()) * multiplier
    return number / 100.0 if is_percent else number


def parse_number(value: Any) -> float | None:
    if is_missing(value):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    raw = unicodedata.normalize("NFKC", str(value)).strip().lower()
    compact = re.sub(r"[^0-9,\.\-+%a-z\u4e00-\u9fff]", "", raw)
    if "," in compact and "." in compact:
        if compact.rfind(",") > compact.rfind("."):
            compact = compact.replace(".", "").replace(",", ".")
        else:
            compact = compact.replace(",", "")
    elif "," in compact:
        parts = compact.split(",")
        compact = ".".join(parts) if len(parts) == 2 and len(parts[1].replace("%", "")) in {1, 2} else "".join(parts)
    is_percent = "%" in compact
    multiplier = 1.0
    for suffix, factor in {"k": 1_000, "m": 1_000_000, "b": 1_000_000_000, "\u4e07": 10_000, "\u4ebf": 100_000_000}.items():
        if suffix in compact:
            multiplier = factor
            break
    match = re.search(r"[-+]?\d*\.?\d+", compact)
    if not match:
        return None
    parsed = float(match.group()) * multiplier
    return parsed / 100.0 if is_percent else parsed


def market_timezone(market: str, reference_date: date | None = None) -> tzinfo:
    try:
        return ZoneInfo(MARKET_TIMEZONES[market])
    except ZoneInfoNotFoundError:
        reference_date = reference_date or datetime.now(timezone.utc).date()
        if market == "US":
            dst_like = 3 <= reference_date.month <= 11
        else:
            dst_like = 3 <= reference_date.month <= 10
        return timezone(timedelta(hours=FALLBACK_MARKET_OFFSETS[market][dst_like]))


def parse_datetime(value: Any, assumed_tz: tzinfo) -> datetime | None:
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day)
    elif isinstance(value, (int, float)) or str(value).strip().isdigit():
        raw = float(value)
        if raw > 10_000_000_000:
            raw /= 1000.0
        try:
            return datetime.fromtimestamp(raw, tz=timezone.utc)
        except (OSError, OverflowError, ValueError):
            return None
    else:
        text = unicodedata.normalize("NFKC", str(value)).strip().replace("Z", "+00:00")
        parsed = None
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
        if parsed is None:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=assumed_tz)
    return parsed


def load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    return [dict(row) for row in csv.DictReader(handle, delimiter=delimiter)]
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Unable to decode delimited file: {path}")
    if suffix == ".jsonl":
        rows = []
        with path.open("r", encoding="utf-8-sig") as handle:
            for line in handle:
                if line.strip():
                    item = json.loads(line)
                    if isinstance(item, dict):
                        rows.append(item)
        return rows
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        return unwrap_records(payload)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX input requires openpyxl; use the bundled workspace Python runtime") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                values = sheet.iter_rows(values_only=True)
                try:
                    headers = next(values)
                except StopIteration:
                    continue
                if not any(not is_missing(header) for header in headers):
                    continue
                names = [clean_text(header) or f"column_{index + 1}" for index, header in enumerate(headers)]
                rows = [dict(zip(names, row)) for row in values if any(not is_missing(cell) for cell in row)]
                if rows:
                    return rows
            return []
        finally:
            workbook.close()
    raise ValueError(f"Unsupported input type: {path.suffix}")


def worksheet_rows(worksheet: Any) -> list[dict[str, Any]]:
    values = worksheet.iter_rows(values_only=True)
    try:
        headers = next(values)
    except StopIteration:
        return []
    names = [clean_text(header) or f"column_{index + 1}" for index, header in enumerate(headers)]
    return [dict(zip(names, row)) for row in values if any(not is_missing(cell) for cell in row)]


def find_structured_sheet(workbook: Any, key: str) -> Any:
    accepted = STRUCTURED_CATALOG_SHEETS[key]
    for worksheet in workbook.worksheets:
        if worksheet.title in accepted:
            return worksheet
    raise ValueError(f"Structured catalog is missing sheet: {key}")


def is_structured_product_catalog(path: Path) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Structured XLSX catalog requires bundled openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        names = set(workbook.sheetnames)
        return all(names & accepted for accepted in STRUCTURED_CATALOG_SHEETS.values())
    finally:
        workbook.close()


def load_structured_product_catalog(path: Path, market: str, allow_out_of_stock: bool = False) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Structured XLSX catalog requires bundled openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        master_rows = worksheet_rows(find_structured_sheet(workbook, "product_master"))
        alias_rows = worksheet_rows(find_structured_sheet(workbook, "identifier_aliases"))
        inventory_rows = worksheet_rows(find_structured_sheet(workbook, "market_inventory"))
    finally:
        workbook.close()

    markets_detected = sorted({normalize_market(row.get("market"), "") for row in master_rows if clean_text(row.get("market"))})
    aliases: dict[str, dict[str, Any]] = {}
    for row in alias_rows:
        row_market = normalize_market(row.get("market"), "")
        sku = clean_text(row.get("sku"))
        if row_market == market and sku:
            aliases[sku] = row

    grouped_inventory: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for row in inventory_rows:
        row_market = normalize_market(row.get("market"), "")
        sku = clean_text(row.get("sku"))
        if row_market != market or not sku:
            continue
        snapshot = clean_text(row.get("snapshot_date")) or "undated"
        grouped_inventory.setdefault(sku, {}).setdefault(snapshot, []).append(row)

    inventory: dict[str, dict[str, Any]] = {}
    snapshot_dates: set[str] = set()
    for sku, snapshots in grouped_inventory.items():
        selected_snapshot = max(snapshots)
        snapshot_dates.add(selected_snapshot)
        rows = snapshots[selected_snapshot]
        inventory[sku] = {
            "stock": sum(parse_number(row.get("available_stock")) or 0.0 for row in rows),
            "in_transit_stock": sum(parse_number(row.get("in_transit_stock")) or 0.0 for row in rows),
            "stock_status": ";".join(sorted({clean_text(row.get("stock_status")) for row in rows if clean_text(row.get("stock_status"))})),
            "inventory_snapshot_date": selected_snapshot,
            "inventory_source_skus": ";".join(sorted({clean_text(row.get("source_sku")) for row in rows if clean_text(row.get("source_sku"))})),
        }

    products: list[dict[str, Any]] = []
    missing_inventory = 0
    out_of_stock = 0
    selected_master_rows = 0
    for source_row in master_rows:
        row_market = normalize_market(source_row.get("market"), "")
        if row_market != market:
            continue
        selected_master_rows += 1
        row = canonicalize_row(source_row, PRODUCT_ALIASES)
        product = {field: clean_text(row.get(field)) for field in PRODUCT_ALIASES}
        sku = clean_text(row.get("sku"))
        if not sku:
            continue
        product["sku"] = sku
        product["market"] = market
        product["title"] = clean_text(row.get("title")) or clean_text(row.get("display_name"))
        product["price"] = parse_number(row.get("promo_price")) or parse_number(row.get("normal_price")) or parse_number(row.get("price"))
        product["margin"] = parse_number(row.get("margin"))
        alias = aliases.get(sku, {})
        product["old_skus"] = clean_text(alias.get("old_skus"))
        product["product_ids"] = clean_text(alias.get("tiktok_product_ids"))
        product["primary_product_id"] = clean_text(alias.get("primary_tiktok_product_id"))
        product["asins"] = clean_text(alias.get("amazon_asins"))
        product["preferred_asin"] = clean_text(alias.get("preferred_asin"))
        stock = inventory.get(sku)
        if stock is None:
            missing_inventory += 1
            if not allow_out_of_stock:
                continue
            product["stock"] = None
        else:
            product.update(stock)
            if stock["stock"] <= 0:
                out_of_stock += 1
                if not allow_out_of_stock:
                    continue
        products.append(product)

    unique = {f"{product['market']}::{product['sku']}": product for product in products}
    metadata = {
        "catalog_mode": "structured_three_table",
        "requested_market": market,
        "markets_detected": markets_detected,
        "selected_master_rows": selected_master_rows,
        "eligible_product_rows": len(unique),
        "missing_inventory_filtered": 0 if allow_out_of_stock else missing_inventory,
        "out_of_stock_filtered": 0 if allow_out_of_stock else out_of_stock,
        "allow_out_of_stock": allow_out_of_stock,
        "inventory_snapshot_dates": sorted(snapshot_dates),
    }
    return list(unique.values()), metadata


def unwrap_records(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("data", "records", "rows", "list", "items", "result", "profiles"):
            if key in payload:
                records = unwrap_records(payload[key])
                if records:
                    return records
        if payload and all(not isinstance(value, (list, dict)) for value in payload.values()):
            return [payload]
    return []


def normalize_market(value: Any, fallback: str) -> str:
    text = clean_text(value).upper()
    if text in {"USA", "UNITED STATES", "美国", "美区"}:
        return "US"
    if text in {"GERMANY", "DEU", "德国", "德区"}:
        return "DE"
    return text or fallback


def extract_video_id(url: str) -> str:
    match = re.search(r"(?:/video/|video_id=|item_id=)(\d{8,})", url)
    if match:
        return match.group(1)
    return "url-" + hashlib.sha256(url.encode("utf-8")).hexdigest()[:16] if url else ""


def extract_creator_name(url: str) -> str:
    match = re.search(r"/@([^/]+)/video/", url)
    return match.group(1) if match else ""


def extract_product_id(url: str) -> str:
    match = re.search(r"/product/(\d{8,})", url)
    return match.group(1) if match else ""


def normalize_creator_handle(value: Any) -> str:
    text = clean_text(value)
    if text.startswith("@"):
        text = text[1:]
    return unicodedata.normalize("NFKC", text).strip().lower()


def extract_profile_handle(url: str) -> str:
    text = clean_text(url)
    match = re.search(r"(?:/tiktok/|/@)([^/?#]+)", text)
    return match.group(1) if match else ""


def normalize_creator_profiles(rows: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    profiles: dict[str, dict[str, Any]] = {}
    for source_row in rows:
        row = canonicalize_row(source_row, CREATOR_PROFILE_ALIASES)
        source_url = clean_text(row.get("creator_profile_source_url"))
        supplied_handle = clean_text(row.get("creator_name"))
        url_handle = extract_profile_handle(source_url)
        handle = normalize_creator_handle(supplied_handle or url_handle)
        if not handle:
            continue
        followers = parse_number(row.get("creator_followers"))
        profile = {
            "creator_name": supplied_handle or url_handle,
            "creator_followers": followers,
            "creator_followers_as_of": clean_text(row.get("creator_followers_as_of")),
            "creator_profile_source": clean_text(row.get("creator_profile_source")) or "external_profile",
            "creator_profile_source_url": source_url,
            "_profile_url_handle": normalize_creator_handle(url_handle),
            "_profile_mismatch": bool(
                supplied_handle
                and url_handle
                and normalize_creator_handle(supplied_handle) != normalize_creator_handle(url_handle)
            ),
        }
        current = profiles.get(handle)
        current_quality = int(bool(current and current.get("creator_followers") is not None)) + int(bool(current and current.get("creator_followers_as_of")))
        new_quality = int(followers is not None) + int(bool(profile["creator_followers_as_of"]))
        if current is None or new_quality >= current_quality:
            profiles[handle] = profile
    return profiles


def apply_creator_size_evidence(rows: list[dict[str, Any]], profiles: dict[str, dict[str, Any]], max_followers: int) -> None:
    for row in rows:
        handle = normalize_creator_handle(row.get("creator_name"))
        profile = profiles.get(handle)
        row["creator_profile_match_basis"] = ""
        if profile:
            for field in ("creator_followers", "creator_followers_as_of", "creator_profile_source", "creator_profile_source_url"):
                row[field] = profile.get(field) if field == "creator_followers" else clean_text(profile.get(field))
            row["creator_profile_match_basis"] = "exact_creator_handle"
            mismatch = bool(profile.get("_profile_mismatch"))
        else:
            mismatch = False
            if row.get("creator_followers") is not None:
                row["creator_followers_as_of"] = clean_text(row.get("creator_followers_as_of") or row.get("captured_at"))
                row["creator_profile_source"] = clean_text(row.get("creator_profile_source") or row.get("source_tier") or "candidate_field")
                row["creator_profile_match_basis"] = "candidate_field"

        followers = row.get("creator_followers")
        source_url = clean_text(row.get("creator_profile_source_url"))
        as_of = clean_text(row.get("creator_followers_as_of"))
        if not handle:
            status, reason = "unverified", "missing_creator_handle"
        elif mismatch:
            status, reason = "profile_mismatch", "profile handle does not exactly match creator"
        elif followers is None:
            status, reason = "unverified", "missing_creator_followers"
        elif not as_of or not clean_text(row.get("creator_profile_source")):
            status, reason = "unverified", "missing_creator_source_or_date"
        elif source_url and extract_profile_handle(source_url) and normalize_creator_handle(extract_profile_handle(source_url)) != handle:
            status, reason = "profile_mismatch", "source URL handle does not exactly match creator"
        elif followers >= max_followers:
            status, reason = "over_limit", f"creator_followers>={max_followers}"
        else:
            status, reason = "verified_small", ""
        row["creator_size_status"] = status
        row["creator_size_rejection_reason"] = reason


def creator_size_eligible(row: dict[str, Any], max_followers: int | None) -> bool:
    if max_followers is None:
        return True
    return row.get("creator_size_status") == "verified_small" and (row.get("creator_followers") or 0) < max_followers


def normalize_candidates(rows: Iterable[dict[str, Any]], market: str, source_tier: str = "") -> list[dict[str, Any]]:
    assumed_tz = market_timezone(market)
    normalized = []
    for source_row in rows:
        row = canonicalize_row(source_row, CANDIDATE_ALIASES)
        result = {field: clean_text(row.get(field)) for field in CANDIDATE_ALIASES}
        for field in NUMERIC_CANDIDATE_FIELDS:
            result[field] = parse_number(row.get(field))
        result["market"] = normalize_market(row.get("market"), market)
        if result["market"] != market:
            continue
        result["video_url"] = clean_text(row.get("video_url"))
        result["video_id"] = clean_text(row.get("video_id")) or extract_video_id(result["video_url"])
        result["creator_name"] = clean_text(row.get("creator_name")) or extract_creator_name(result["video_url"])
        result["product_url"] = clean_text(row.get("product_url"))
        result["product_id"] = clean_text(row.get("product_id")) or extract_product_id(result["product_url"])
        if not result["currency"]:
            result["currency"] = {"US": "USD", "DE": "EUR"}.get(market, "")
        published = parse_datetime(row.get("published_at"), assumed_tz)
        captured = parse_datetime(row.get("captured_at"), assumed_tz) or datetime.now(timezone.utc)
        result["published_at"] = published.isoformat() if published else ""
        result["captured_at"] = captured.isoformat()
        interaction_values = [result.get(field) for field in ("likes", "comments", "shares")]
        if result["engagement_rate"] is None and result["views"] and result["views"] > 0 and any(value is not None for value in interaction_values):
            interactions = sum(value or 0 for value in interaction_values)
            result["engagement_rate"] = interactions / result["views"]
        result["source_tier"] = result["source_tier"] or source_tier or "unknown"
        result["evidence_status"] = result["evidence_status"] or "available"
        result["_key"] = result["video_id"] or result["video_url"]
        normalized.append(result)
    deduped: dict[str, dict[str, Any]] = {}
    for row in normalized:
        key = row["_key"]
        if not key:
            continue
        if key not in deduped or evidence_count(row) > evidence_count(deduped[key]):
            deduped[key] = row
    return list(deduped.values())


def normalize_products(rows: Iterable[dict[str, Any]], market: str, allow_out_of_stock: bool = False) -> list[dict[str, Any]]:
    products = []
    for source_row in rows:
        row = canonicalize_row(source_row, PRODUCT_ALIASES)
        product = {field: clean_text(row.get(field)) for field in PRODUCT_ALIASES}
        product["market"] = normalize_market(row.get("market"), market)
        for field in ("price", "stock", "margin"):
            product[field] = parse_number(row.get(field))
        if product["sku"] and (allow_out_of_stock or (product.get("stock") is not None and product["stock"] > 0)):
            products.append(product)
    unique: dict[str, dict[str, Any]] = {}
    for product in products:
        unique[f"{product['market']}::{product['sku']}"] = product
    return list(unique.values())


def evidence_count(row: dict[str, Any]) -> int:
    return sum(not is_missing(row.get(field)) for field in CANDIDATE_ALIASES)


def percentile_map(rows: list[dict[str, Any]], field: str) -> dict[str, float]:
    values = sorted((float(row[field]), row["_key"]) for row in rows if row.get(field) is not None)
    if not values:
        return {}
    if len(values) == 1:
        return {values[0][1]: 100.0}
    result: dict[str, float] = {}
    index = 0
    while index < len(values):
        end = index
        while end + 1 < len(values) and values[end + 1][0] == values[index][0]:
            end += 1
        percentile = 100.0 * ((index + end) / 2.0) / (len(values) - 1)
        for _, key in values[index:end + 1]:
            result[key] = percentile
        index = end + 1
    return result


def weighted_available(parts: list[tuple[float | None, float]]) -> float:
    available = [(value, weight) for value, weight in parts if value is not None]
    if not available:
        return 0.0
    total_weight = sum(weight for _, weight in available)
    return sum(float(value) * weight for value, weight in available) / total_weight


def tokenize(value: Any) -> set[str]:
    text = unicodedata.normalize("NFKC", clean_text(value)).lower()
    latin = set(re.findall(r"[a-z0-9]{2,}", text))
    chinese_chunks = re.findall(r"[\u4e00-\u9fff]+", text)
    cjk = set()
    for chunk in chinese_chunks:
        cjk.add(chunk)
        cjk.update(chunk[index:index + 2] for index in range(max(0, len(chunk) - 1)))
    return latin | cjk


def similarity(left: Any, right: Any) -> float:
    a, b = tokenize(left), tokenize(right)
    if not a or not b:
        return 0.0
    return len(a & b) / math.sqrt(len(a) * len(b))


def identifier_tokens(value: Any) -> set[str]:
    return {
        token.upper()
        for token in re.split(r"[;,|\s]+", clean_text(value))
        if token and len(token) >= 3
    }


def build_product_identifier_index(products: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        identifiers = (
            identifier_tokens(product.get("product_ids"))
            | identifier_tokens(product.get("asins"))
            | identifier_tokens(product.get("old_skus"))
        )
        sku = clean_text(product.get("sku")).upper()
        if sku:
            identifiers.add(sku)
        for identifier in identifiers:
            index[identifier].append(product)
    return dict(index)


def infer_angle(row: dict[str, Any]) -> str:
    if clean_text(row.get("content_angle")):
        return clean_text(row["content_angle"])
    text = " ".join(clean_text(row.get(field)).lower() for field in ("caption", "transcript", "hook", "pain_point", "proof_action"))
    best = (0, "product-demo")
    for angle, words in ANGLE_KEYWORDS.items():
        count = sum(word in text for word in words)
        if count > best[0]:
            best = (count, angle)
    return best[1]


def product_match(
    candidate: dict[str, Any],
    products: list[dict[str, Any]],
    identifier_index: dict[str, list[dict[str, Any]]] | None = None,
    identifier_match_only: bool = False,
) -> tuple[dict[str, Any] | None, float, list[str]]:
    best_product, best_score, best_reasons = None, 0.0, []
    if identifier_match_only and clean_text(candidate.get("catalog_match_status")) in {"ambiguous_exact_id", "exact_not_stock_positive"}:
        return None, 0.0, ["source identifier is ambiguous or not stock-positive in the market catalog"]
    candidate_product = " ".join(clean_text(candidate.get(field)) for field in ("product_title", "category", "product_id"))
    candidate_story = " ".join(clean_text(candidate.get(field)) for field in ("caption", "transcript", "pain_point", "hook"))
    candidate_proof = " ".join(clean_text(candidate.get(field)) for field in ("proof_action", "transcript", "caption"))
    candidate_identifiers = identifier_tokens(candidate.get("product_id"))
    candidate_identifiers.update(re.findall(r"\b[A-Z0-9]{10,}\b", clean_text(candidate.get("product_url")).upper()))
    if identifier_match_only:
        lookup = identifier_index or build_product_identifier_index(products)
        exact_products: dict[str, dict[str, Any]] = {}
        for identifier in candidate_identifiers:
            for product in lookup.get(identifier, []):
                exact_products[f"{product.get('market')}::{product.get('sku')}"] = product
        if not exact_products:
            return None, 0.0, ["no exact market catalog identifier match"]
        if len(exact_products) > 1:
            return None, 0.0, ["source identifier maps to multiple eligible market SKUs"]
        products = list(exact_products.values())
    for product in products:
        if product["market"] not in {"", candidate["market"]}:
            continue
        type_score = similarity(candidate_product, " ".join(str(product.get(field) or "") for field in ("title", "category", "product_type"))) * 25
        pain_score = similarity(candidate_story, " ".join(str(product.get(field) or "") for field in ("pain_points", "selling_points"))) * 25
        proof_score = similarity(candidate_proof, " ".join(str(product.get(field) or "") for field in ("proof_actions", "selling_points"))) * 15
        attr_score = similarity(candidate_product + " " + candidate_story, " ".join(str(product.get(field) or "") for field in ("fit", "colors", "neckline", "sleeve", "fabric"))) * 15
        market_score = 7.5 if product["market"] == candidate["market"] else 0.0
        market_score += similarity(candidate_story, product.get("scenarios")) * 2.5
        price_score = 0.0
        if candidate.get("price") is not None and product.get("price") is not None and product["price"] > 0:
            ratio = abs(candidate["price"] - product["price"]) / product["price"]
            price_score = 5.0 if ratio <= 0.30 else 2.5 if ratio <= 0.60 else 0.0
        readiness = (2.5 if product.get("stock") is not None and product["stock"] > 0 else 0.0) + (2.5 if product.get("margin") is not None and product["margin"] > 0 else 0.0)
        exact = False
        exact_identifier = False
        sku = clean_text(product.get("sku")).lower()
        if sku and sku in candidate_product.lower():
            exact = True
        product_identifiers = identifier_tokens(product.get("product_ids")) | identifier_tokens(product.get("asins")) | identifier_tokens(product.get("old_skus"))
        product_identifiers.add(clean_text(product.get("sku")).upper())
        if candidate_identifiers & product_identifiers:
            exact = True
            exact_identifier = True
        score = type_score + pain_score + proof_score + attr_score + market_score + price_score + readiness
        if exact:
            score = max(score, 95.0 if exact_identifier else 85.0)
        score = min(100.0, score)
        if score > best_score:
            reasons = []
            if exact_identifier:
                reasons.append("source product ID or ASIN matches the market catalog")
            if exact:
                if not exact_identifier:
                    reasons.append("source product references the SKU")
            if type_score >= 8:
                reasons.append("product type/category overlap")
            if pain_score >= 6:
                reasons.append("buyer-pain overlap")
            if proof_score >= 4:
                reasons.append("proof action is compatible")
            if attr_score >= 4:
                reasons.append("garment attributes overlap")
            if readiness:
                reasons.append("stock or margin is production-ready")
            best_product, best_score, best_reasons = product, score, reasons or ["weak text match; requires semantic review"]
    return best_product, round(best_score, 2), best_reasons


def score_candidates(
    rows: list[dict[str, Any]],
    products: list[dict[str, Any]],
    analysis_date: date,
    identifier_match_only: bool = False,
) -> None:
    rank_fields = {field: percentile_map(rows, field) for field in ("views", "engagement_rate", "sales", "gmv", "view_growth", "sales_growth")}
    identifier_index = build_product_identifier_index(products) if identifier_match_only else None
    for row in rows:
        growth_values = [rank_fields[field].get(row["_key"]) for field in ("view_growth", "sales_growth")]
        growth = max((value for value in growth_values if value is not None), default=None)
        assumed_tz = market_timezone(row["market"], analysis_date)
        published = parse_datetime(row.get("published_at"), assumed_tz)
        recency = None
        if published:
            age = (analysis_date - published.astimezone(assumed_tz).date()).days
            recency = 100.0 if age in {0, 1} else 80.0 if 0 <= age <= 7 else 50.0 if age <= 30 else 20.0
        row["hot_score"] = round(weighted_available([
            (rank_fields["sales"].get(row["_key"]), 40),
            (rank_fields["gmv"].get(row["_key"]), 30),
            (rank_fields["views"].get(row["_key"]), 10),
            (rank_fields["engagement_rate"].get(row["_key"]), 5),
            (growth, 10),
            (recency, 5),
        ]), 2)
        replicability = 35.0
        replicability += 12 if row.get("caption") else 0
        replicability += 15 if row.get("transcript") else 0
        replicability += 12 if row.get("hook") else 0
        replicability += 14 if row.get("proof_action") else 0
        replicability += 7 if row.get("cta") else 0
        duration = row.get("duration_sec")
        replicability += 5 if duration is not None and duration <= 60 else 0
        replicability -= 20 if row.get("risk_note") else 0
        if row.get("evidence_status") == "limited_evidence":
            replicability = min(replicability, 60.0)
        row["replicability_score"] = round(max(0.0, min(100.0, replicability)), 2)
        present = [bool(row.get("video_id") or row.get("video_url")), bool(row.get("creator_name")), bool(row.get("published_at")), bool(row.get("market")), bool(row.get("product_title")), row.get("views") is not None, row.get("sales") is not None or row.get("gmv") is not None]
        tier_factor = {"backend_verified": 1.0, "official_export": 0.9, "page_extraction": 0.75}.get(row.get("source_tier"), 0.6)
        confidence = 100.0 * sum(present) / len(present) * tier_factor
        if row.get("evidence_status") == "limited_evidence":
            confidence = min(confidence, 60.0)
        row["evidence_confidence"] = round(confidence, 2)
        product, match_score, reasons = product_match(row, products, identifier_index, identifier_match_only)
        row["matched_sku"] = product["sku"] if product else ""
        row["matched_product"] = product["title"] if product else ""
        row["product_match_score"] = match_score
        row["match_reasons"] = reasons
        row["_product"] = product or {}
        row["content_angle"] = infer_angle(row)
        row["final_score"] = round(0.55 * row["hot_score"] + 0.30 * match_score + 0.10 * row["replicability_score"] + 0.05 * row["evidence_confidence"], 2)


def selection_sort_key(row: dict[str, Any], rank_by: str) -> tuple[float, float, float, str]:
    if rank_by == "sales":
        return (-(row.get("sales") if row.get("sales") is not None else -1), -(row.get("gmv") if row.get("gmv") is not None else -1), -row["final_score"], row["_key"])
    if rank_by == "gmv":
        return (-(row.get("gmv") if row.get("gmv") is not None else -1), -(row.get("sales") if row.get("sales") is not None else -1), -row["final_score"], row["_key"])
    return (-row["final_score"], -row["hot_score"], -(row.get("sales") if row.get("sales") is not None else -1), row["_key"])


def select_diverse(
    rows: list[dict[str, Any]],
    top: int,
    min_match: float,
    max_per_sku: int,
    max_per_creator: int,
    max_creator_followers: int | None = None,
    rank_by: str = "final",
    exclude_ids: set[str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    exclude_ids = exclude_ids or set()
    ranked = sorted(rows, key=lambda row: selection_sort_key(row, rank_by))
    eligible = [
        row for row in ranked
        if row.get("matched_sku")
        and row["product_match_score"] >= min_match
        and creator_size_eligible(row, max_creator_followers)
        and row["_key"] not in exclude_ids
    ]
    selected: list[dict[str, Any]] = []
    sku_counts: Counter[str] = Counter()
    creator_counts: Counter[str] = Counter()

    def can_add(row: dict[str, Any]) -> bool:
        return sku_counts[row["matched_sku"]] < max_per_sku and creator_counts[row.get("creator_name") or "unknown"] < max_per_creator

    for angle in dict.fromkeys(row["content_angle"] for row in eligible):
        if len({item["content_angle"] for item in selected}) >= 4 or len(selected) >= top:
            break
        candidate = next((row for row in eligible if row["content_angle"] == angle and row not in selected and can_add(row)), None)
        if candidate:
            selected.append(candidate)
            sku_counts[candidate["matched_sku"]] += 1
            creator_counts[candidate.get("creator_name") or "unknown"] += 1
    for row in eligible:
        if len(selected) >= top:
            break
        if row not in selected and can_add(row):
            selected.append(row)
            sku_counts[row["matched_sku"]] += 1
            creator_counts[row.get("creator_name") or "unknown"] += 1
    selected.sort(key=lambda row: selection_sort_key(row, rank_by))
    watchlist = [row for row in ranked if row not in selected][:20]
    return selected, watchlist


def selection_exclusion_clue(row: dict[str, Any]) -> str:
    if row.get("creator_size_status") and row.get("creator_size_status") != "verified_small":
        return clean_text(row.get("creator_size_rejection_reason")) or clean_text(row.get("creator_size_status"))
    if not row.get("matched_sku"):
        return "no qualified SKU match"
    return "diversity, duplicate-exclusion, or selected-list limit"


def core_missingness(rows: list[dict[str, Any]]) -> float:
    if not rows:
        return 1.0
    missing = 0
    for row in rows:
        checks = [
            bool(row.get("video_id") or row.get("video_url")),
            bool(row.get("creator_name")), bool(row.get("published_at")), bool(row.get("market")),
            bool(row.get("product_title")), row.get("views") is not None,
            row.get("sales") is not None or row.get("gmv") is not None,
        ]
        missing += sum(not check for check in checks)
    return missing / (len(rows) * len(CORE_FIELDS))


def parse_hashtags(value: Any) -> list[str]:
    text = clean_text(value).replace("#Imily Bela", "#Imily_Bela")
    return [tag.replace("#Imily_Bela", "#Imily Bela") for tag in re.findall(r"#[^\s,;]+", text)]


def replication_package_complete(row: dict[str, Any]) -> bool:
    product = row.get("_product") or {}
    pain = clean_text(row.get("pain_point") or product.get("pain_points"))
    proof = clean_text(row.get("proof_action") or product.get("proof_actions"))
    source_structure = clean_text(row.get("source_video_structure"))
    source_voiceover = clean_text(row.get("source_voiceover_full") or row.get("voiceover") or row.get("transcript"))
    source_voiceover_zh = clean_text(row.get("source_voiceover_zh") or row.get("voiceover_zh"))
    hashtags = parse_hashtags(row.get("hashtags"))
    return all([
        clean_text(row.get("hook")),
        pain,
        proof,
        clean_text(row.get("cta")),
        source_structure,
        source_voiceover,
        source_voiceover_zh,
        clean_text(row.get("cover_copy")),
        len(hashtags) == 5,
        "#Imily Bela" in hashtags,
    ])


def serializable_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if not key.startswith("_")}


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = [*CANDIDATE_ALIASES.keys(), *CREATOR_OUTPUT_FIELDS, "hot_score", "product_match_score", "replicability_score", "evidence_confidence", "final_score", "matched_sku", "matched_product", "match_reasons"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            output = serializable_row(row)
            output["match_reasons"] = "; ".join(output.get("match_reasons", []))
            writer.writerow(output)


def md(value: Any) -> str:
    return clean_text(value).replace("|", "\\|") or "-"


def report_markdown(manifest: dict[str, Any], selected: list[dict[str, Any]], watchlist: list[dict[str, Any]]) -> str:
    status = "PASS" if manifest["passed"] else "NEEDS ATTENTION"
    store_suffix = f" {manifest['store_id']}" if manifest.get("store_id") else ""
    lines = [
        f"# FastMoss Daily Viral Replication Report - {manifest['analysis_date']} {manifest['market']}{store_suffix}",
        "",
        f"Status: **{status}**",
        "",
        f"Candidates: {manifest['candidate_count']} | Products: {manifest['product_count']} | Selected: {manifest['selected_count']} | Rank by: {manifest['rank_by']} | Core missingness: {manifest['core_missingness_pct']}%",
        "",
        "## Acceptance Gates",
        "",
        "| Gate | Result |",
        "| --- | --- |",
    ]
    for name, passed in manifest["gates"].items():
        lines.append(f"| {name} | {'PASS' if passed else 'FAIL'} |")
    lines.extend(["", "## Selected Recommendations", "", "| Rank | Video | Creator | Followers | Sales | GMV | Hot | Match | Final | SKU | Angle |", "| ---: | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | --- | --- |"])
    for index, row in enumerate(selected, 1):
        link = f"[{md(row.get('video_id'))}]({row.get('video_url')})" if row.get("video_url") else md(row.get("video_id"))
        followers = f"{int(row['creator_followers']):,}" if row.get("creator_followers") is not None else "-"
        lines.append(f"| {index} | {link} | {md(row.get('creator_name'))} | {followers} | {md(row.get('sales'))} | {md(row.get('gmv'))} | {row['hot_score']:.1f} | {row['product_match_score']:.1f} | {row['final_score']:.1f} | {md(row.get('matched_sku'))} | {md(row.get('content_angle'))} |")
    lines.extend(["", "## Replication Battle Cards", ""])
    for index, row in enumerate(selected, 1):
        product = row.get("_product") or {}
        source_structure = row.get("source_video_structure") or "[actual source-video structure required]"
        source_voiceover = row.get("source_voiceover_full") or row.get("voiceover") or row.get("transcript") or "[complete source-video voiceover required]"
        source_voiceover_zh = row.get("source_voiceover_zh") or row.get("voiceover_zh") or "[需完成完整口播中文翻译]"
        hashtags = row.get("hashtags") or "[exactly five hashtags required; include #Imily Bela]"
        lines.extend([
            f"### {index}. {md(row.get('matched_sku'))} - {md(row.get('matched_product'))}",
            "",
            f"- Source: {row.get('video_url') or '-'}",
            f"- Creator / published: {md(row.get('creator_name'))} / {md(row.get('published_at'))}",
            f"- Creator followers: {md(row.get('creator_followers'))}",
            f"- Evidence: {md(row.get('source_tier'))}; {md(row.get('evidence_status'))}; confidence {row['evidence_confidence']:.1f}",
            f"- Why viral: hot {row['hot_score']:.1f}; views {md(row.get('views'))}; sales {md(row.get('sales'))}; GMV {md(row.get('gmv'))}",
            f"- Similar replicable SKU: {md(row.get('matched_sku'))}; match {row['product_match_score']:.1f}; {'; '.join(row.get('match_reasons') or [])}",
            f"- Buyer pain: {md(row.get('pain_point') or product.get('pain_points'))}",
            f"- Visible proof: {md(row.get('proof_action') or product.get('proof_actions'))}",
            f"- Actual source-video structure: {md(source_structure)}",
            f"- Full source voiceover/transcript: {md(source_voiceover)}",
            f"- Chinese translation: {md(source_voiceover_zh)}",
            f"- Cover copy: {md(row.get('cover_copy'))}",
            f"- Hashtags: {md(hashtags)}",
            "",
        ])
    lines.extend(["## Watchlist", "", "| Video | Score | Exclusion clue |", "| --- | ---: | --- |"])
    for row in watchlist[:10]:
        clue = selection_exclusion_clue(row)
        lines.append(f"| {md(row.get('video_id') or row.get('video_url'))} | {row['final_score']:.1f} | {clue} |")
    return "\n".join(lines) + "\n"


def build_manifest(args: argparse.Namespace, candidates: list[dict[str, Any]], products: list[dict[str, Any]], selected: list[dict[str, Any]], catalog_metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    missingness = core_missingness(candidates)
    top30 = sorted(candidates, key=lambda row: (-row["final_score"], row["_key"]))[:30]
    decomposition_ready = sum(bool(row.get("caption") or row.get("transcript")) and bool(row.get("product_title")) for row in top30)
    ready_ratio = decomposition_ready / len(top30) if top30 else 0.0
    angles = len({row["content_angle"] for row in selected})
    replication_ready = sum(replication_package_complete(row) for row in selected)
    creator_size_statuses = dict(Counter(row.get("creator_size_status") or "not_checked" for row in candidates))
    creator_size_verified = (
        not args.creator_profiles
        or bool(selected) and all(row.get("creator_size_status") == "verified_small" and (row.get("creator_followers") or 0) < args.max_creator_followers for row in selected)
    )
    inventory_age_days = None
    inventory_fresh = True
    if catalog_metadata and catalog_metadata.get("catalog_mode") == "structured_three_table":
        snapshots = [date.fromisoformat(value) for value in catalog_metadata.get("inventory_snapshot_dates", []) if value]
        if snapshots:
            inventory_age_days = (date.fromisoformat(args.analysis_date) - max(snapshots)).days
            inventory_fresh = 0 <= inventory_age_days <= 1
        else:
            inventory_fresh = False
    gates = {
        "candidate_floor": len(candidates) >= args.min_candidates,
        "core_missingness_lte_20pct": missingness <= 0.20,
        "top30_decomposition_ready_gte_60pct": ready_ratio >= 0.60,
        "selected_count": len(selected) == args.top,
        "all_selected_have_sku": bool(selected) and all(row.get("matched_sku") for row in selected),
        "four_content_angles": len(selected) < 4 or angles >= 4,
        "catalog_market_isolated": bool(products) and all(product.get("market") == args.market for product in products),
        "selected_have_market_stock": catalog_metadata is None or catalog_metadata.get("catalog_mode") != "structured_three_table" or all((row.get("_product") or {}).get("stock", 0) > 0 for row in selected),
        "catalog_inventory_fresh": inventory_fresh,
        "selected_creator_size_verified": creator_size_verified,
        "selected_replication_package_complete": bool(selected) and replication_ready == len(selected),
    }
    return {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "analysis_date": args.analysis_date,
        "market": args.market,
        "store_id": args.store_id,
        "candidate_input": str(Path(args.candidates).resolve()),
        "candidate_sha256": file_sha256(Path(args.candidates)),
        "product_input": str(Path(args.products).resolve()),
        "product_sha256": file_sha256(Path(args.products)),
        "candidate_count": len(candidates),
        "product_count": len(products),
        "selected_count": len(selected),
        "rank_by": args.rank_by,
        "excluded_video_count": args.excluded_video_count,
        "source_tiers": dict(Counter(row.get("source_tier") or "unknown" for row in candidates)),
        "creator_profile_input": str(Path(args.creator_profiles).resolve()) if args.creator_profiles else "",
        "creator_profile_sha256": file_sha256(Path(args.creator_profiles)) if args.creator_profiles else "",
        "max_creator_followers": args.max_creator_followers if args.creator_profiles else None,
        "creator_size_statuses": creator_size_statuses,
        "identifier_match_only": args.identifier_match_only,
        "core_missingness_pct": round(missingness * 100, 2),
        "top30_decomposition_ready_pct": round(ready_ratio * 100, 2),
        "content_angle_count": angles,
        "replication_ready_selected_count": replication_ready,
        "inventory_age_days": inventory_age_days,
        "catalog": catalog_metadata or {"catalog_mode": "legacy_single_table", "requested_market": args.market},
        "gates": gates,
        "passed": all(gates.values()),
    }


def failure_events(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    code_map = {
        "candidate_floor": "candidate_floor_not_met",
        "core_missingness_lte_20pct": "field_missing",
        "top30_decomposition_ready_gte_60pct": "field_missing",
        "selected_count": "candidate_floor_not_met",
        "all_selected_have_sku": "product_catalog_incomplete",
        "four_content_angles": "field_missing",
        "catalog_market_isolated": "product_catalog_incomplete",
        "selected_have_market_stock": "product_catalog_incomplete",
        "catalog_inventory_fresh": "product_catalog_incomplete",
        "selected_creator_size_verified": "creator_size_not_qualified",
        "selected_replication_package_complete": "field_missing",
    }
    events = []
    for gate, passed in manifest["gates"].items():
        if not passed:
            events.append({
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "stage": "acceptance",
                "severity": "error",
                "code": code_map[gate],
                "message": f"Acceptance gate failed: {gate}",
                "attempted_tier": ",".join(manifest["source_tiers"].keys()),
                "fallback": "enrich inputs and rerun",
            })
    return events


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidates", required=True)
    parser.add_argument("--products", required=True)
    parser.add_argument("--market", required=True, choices=sorted(MARKET_TIMEZONES))
    parser.add_argument("--store-id", default="", help="Store/shop identifier for per-store reporting")
    parser.add_argument("--source-tier", choices=["backend_verified", "official_export", "page_extraction"], default="")
    parser.add_argument("--analysis-date", help="YYYY-MM-DD; defaults to previous market day")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--top", type=int, default=10)
    parser.add_argument("--rank-by", choices=["final", "sales", "gmv"], default="final", help="Primary final-selection ordering")
    parser.add_argument("--min-candidates", type=int, default=100)
    parser.add_argument("--min-match", type=float, default=20.0)
    parser.add_argument("--identifier-match-only", action="store_true", help="Bind SKUs only through exact market-catalog identifiers")
    parser.add_argument("--max-per-sku", type=int, default=2)
    parser.add_argument("--max-per-creator", type=int, default=2)
    parser.add_argument("--creator-profiles", default="", help="Optional CSV/JSON/XLSX creator follower evidence keyed by exact handle")
    parser.add_argument("--max-creator-followers", type=int, default=50000, help="Maximum creator followers allowed in final selection when creator profiles are supplied")
    parser.add_argument("--exclude-video-ids", default="", help="Optional newline, CSV, or JSON file of video IDs/URLs to exclude from final selection")
    parser.add_argument("--allow-out-of-stock", action="store_true", help="Include zero-stock or missing-inventory products from a structured catalog")
    args = parser.parse_args(argv)
    if not args.analysis_date:
        args.analysis_date = (datetime.now(market_timezone(args.market)).date() - timedelta(days=1)).isoformat()
    date.fromisoformat(args.analysis_date)
    if args.max_creator_followers <= 0:
        raise ValueError("--max-creator-followers must be positive")
    args.excluded_video_count = 0
    return args


def load_excluded_video_ids(path: Path) -> set[str]:
    if not path.is_file():
        raise FileNotFoundError("Excluded video ID input must be an existing file")
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            values = payload.get("video_ids") or payload.get("videos") or payload.get("ids") or []
        else:
            values = payload
        ids = set()
        for value in values:
            if isinstance(value, dict):
                normalized = canonicalize_row(value, CANDIDATE_ALIASES)
                text_value = clean_text(normalized.get("video_id") or normalized.get("video_url"))
            else:
                text_value = clean_text(value)
            if text_value:
                ids.add(text_value)
        return ids
    if path.suffix.lower() not in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        return {line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip()}
    rows = load_rows(path)
    if rows:
        ids = set()
        for row in rows:
            normalized = canonicalize_row(row, CANDIDATE_ALIASES)
            value = clean_text(normalized.get("video_id") or normalized.get("video_url"))
            if value:
                ids.add(value)
        return ids
    return set()


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    candidate_path, product_path = Path(args.candidates), Path(args.products)
    if not candidate_path.is_file() or not product_path.is_file():
        raise FileNotFoundError("Candidate and product inputs must be existing files")
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    candidates = normalize_candidates(load_rows(candidate_path), args.market, args.source_tier)
    creator_profiles = {}
    if args.creator_profiles:
        profile_path = Path(args.creator_profiles)
        if not profile_path.is_file():
            raise FileNotFoundError("Creator profile evidence input must be an existing file")
        creator_profiles = normalize_creator_profiles(load_rows(profile_path))
        apply_creator_size_evidence(candidates, creator_profiles, args.max_creator_followers)
    catalog_metadata: dict[str, Any] | None = None
    if is_structured_product_catalog(product_path):
        products, catalog_metadata = load_structured_product_catalog(product_path, args.market, args.allow_out_of_stock)
    else:
        products = normalize_products(load_rows(product_path), args.market, args.allow_out_of_stock)
    if not candidates:
        raise ValueError("No candidate rows could be normalized")
    if not products:
        raise ValueError("No products with concrete SKUs could be normalized")
    score_candidates(candidates, products, date.fromisoformat(args.analysis_date), args.identifier_match_only)
    ranked = sorted(candidates, key=lambda row: selection_sort_key(row, args.rank_by))
    exclude_ids = load_excluded_video_ids(Path(args.exclude_video_ids)) if args.exclude_video_ids else set()
    args.excluded_video_count = len(exclude_ids)
    selected, watchlist = select_diverse(
        ranked,
        args.top,
        args.min_match,
        args.max_per_sku,
        args.max_per_creator,
        args.max_creator_followers if args.creator_profiles else None,
        args.rank_by,
        exclude_ids,
    )
    manifest = build_manifest(args, ranked, products, selected, catalog_metadata)
    write_csv(output_dir / "normalized-candidates.csv", ranked)
    write_json(output_dir / "ranked-candidates.json", [serializable_row(row) for row in ranked])
    write_json(output_dir / "selected.json", [serializable_row(row) for row in selected])
    write_json(output_dir / "top10.json", [serializable_row(row) for row in selected])
    write_json(output_dir / "run-manifest.json", manifest)
    (output_dir / "daily-report.md").write_text(report_markdown(manifest, selected, watchlist), encoding="utf-8")
    with (output_dir / "failure-log.jsonl").open("w", encoding="utf-8") as handle:
        for event in failure_events(manifest):
            handle.write(json.dumps(event, ensure_ascii=False) + "\n")
    print(json.dumps({"passed": manifest["passed"], "output_dir": str(output_dir.resolve()), "selected": len(selected), "candidates": len(candidates)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(json.dumps({"error": type(exc).__name__, "message": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise
