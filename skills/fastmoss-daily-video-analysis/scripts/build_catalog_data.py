#!/usr/bin/env python3
"""Build normalized US/DE product catalog tables from the user's source workbooks."""

from __future__ import annotations

import argparse
import json
import math
import posixpath
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


STORE_COLUMNS = {
    11: "store_32",
    12: "store_4",
    13: "local_store",
    14: "store_3",
    15: "store_35",
    16: "store_20",
    17: "store_16",
}
IMAGE_COLUMNS = {4: "main", 5: "white", 6: "scene"}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    normalized = re.sub(r"\s+", " ", str(value)).strip()
    if normalized.upper() in {"#N/A", "#REF!", "#VALUE!", "#NAME?", "#DIV/0!"}:
        return ""
    return normalized


def identifier(value: Any) -> str:
    return text(value).upper()


def number(value: Any) -> float | None:
    if value is None or text(value) in {"", "-", "#REF!", "#VALUE!", "#N/A"}:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    raw = text(value).replace(" ", "").replace("$", "").replace("EUR", "").replace("USD", "")
    raw = raw.replace(chr(8364), "").replace(chr(165), "").replace(chr(65509), "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    raw = raw.replace("%", "")
    try:
        parsed = float(raw)
    except ValueError:
        return None
    return parsed / 100.0 if "%" in text(value) else parsed


def iso_mtime(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()


def snapshot_date(path: Path) -> str:
    match = re.search(r"(20\d{2})[-_.](\d{1,2})[-_.](\d{1,2})", path.name)
    if not match:
        return ""
    year, month, day = (int(part) for part in match.groups())
    return f"{year:04d}-{month:02d}-{day:02d}"


class AliasResolver:
    def __init__(self) -> None:
        self.targets: dict[str, set[str]] = defaultdict(set)

    def add(self, current: Any, old: Any) -> None:
        current_id, old_id = identifier(current), identifier(old)
        if current_id and old_id and current_id != old_id:
            self.targets[old_id].add(current_id)

    def resolve(self, value: Any) -> tuple[str, bool]:
        start = identifier(value)
        if not start:
            return "", False
        current, seen = start, {start}
        while current in self.targets:
            targets = self.targets[current]
            if len(targets) != 1:
                return start, True
            next_value = next(iter(targets))
            if next_value in seen:
                return start, True
            current = next_value
            seen.add(current)
        return current, False


def parse_image_anchors(path: Path) -> dict[str, dict[int, Counter[str]]]:
    result: dict[str, dict[int, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
    namespaces = {
        "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "x": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    with zipfile.ZipFile(path) as archive:
        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        workbook_targets = {
            rel.attrib["Id"]: posixpath.normpath(posixpath.join("xl", rel.attrib["Target"]))
            for rel in workbook_rels
        }
        sheets = workbook_xml.find("m:sheets", namespaces)
        if sheets is None:
            return result
        for sheet in sheets:
            sheet_name = sheet.attrib["name"]
            sheet_path = workbook_targets[sheet.attrib[f"{{{namespaces['r']}}}id"]]
            sheet_xml = ET.fromstring(archive.read(sheet_path))
            drawing = sheet_xml.find("m:drawing", namespaces)
            if drawing is None:
                continue
            rel_path = posixpath.join(posixpath.dirname(sheet_path), "_rels", posixpath.basename(sheet_path) + ".rels")
            sheet_rels = ET.fromstring(archive.read(rel_path))
            sheet_targets = {
                rel.attrib["Id"]: posixpath.normpath(posixpath.join(posixpath.dirname(sheet_path), rel.attrib["Target"]))
                for rel in sheet_rels
            }
            drawing_path = sheet_targets[drawing.attrib[f"{{{namespaces['r']}}}id"]]
            drawing_xml = ET.fromstring(archive.read(drawing_path))
            for anchor in drawing_xml:
                start = anchor.find("x:from", namespaces)
                blip = anchor.find(".//a:blip", namespaces)
                if start is None or blip is None:
                    continue
                column = int(start.find("x:col", namespaces).text) + 1
                row = int(start.find("x:row", namespaces).text) + 1
                image_kind = IMAGE_COLUMNS.get(column)
                if image_kind:
                    result[sheet_name][row][image_kind] += 1
    return result


def load_tiktok_styles(path: Path, resolver: AliasResolver) -> list[dict[str, Any]]:
    anchors = parse_image_anchors(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    records: list[dict[str, Any]] = []
    try:
        for segment, worksheet in (("standard_product", workbook.worksheets[1]), ("apparel", workbook.worksheets[2])):
            for row_number, row in enumerate(worksheet.iter_rows(min_row=2, max_col=34, values_only=True), start=2):
                current = identifier(row[1])
                if not current:
                    continue
                old = identifier(row[2])
                resolver.add(current, old)
                product_ids = []
                product_id_sources = []
                for index, store in STORE_COLUMNS.items():
                    product_id = identifier(row[index])
                    if product_id:
                        product_ids.append(product_id)
                        product_id_sources.append(f"{store}:{product_id}")
                store_names = {part.strip() for part in text(row[10]).split(",") if part.strip()}
                primary_product_id = ""
                store_preference = {
                    "32\u5e97": 11,
                    "4\u5e97": 12,
                    "\u672c\u571f\u5e97": 13,
                    "3\u5e97": 14,
                    "35\u5e97": 15,
                    "20\u5e97": 16,
                    "16\u5e97": 17,
                }
                for store_name in store_names:
                    index = store_preference.get(store_name)
                    if index is not None and identifier(row[index]):
                        primary_product_id = identifier(row[index])
                        break
                primary_product_id = primary_product_id or (product_ids[0] if product_ids else "")
                image_counts = anchors.get(worksheet.title, {}).get(row_number, Counter())
                records.append({
                    "segment": segment,
                    "source_sheet": worksheet.title,
                    "source_row": row_number,
                    "current": current,
                    "old": old,
                    "planning_type": text(row[6]),
                    "style_label": text(row[7]),
                    "normal_price": number(row[8]),
                    "promo_price": number(row[9]),
                    "stores": text(row[10]),
                    "product_ids": sorted(set(product_ids)),
                    "product_id_sources": sorted(set(product_id_sources)),
                    "primary_product_id": primary_product_id,
                    "season": text(row[18]),
                    "owner": text(row[19]),
                    "team": text(row[20]),
                    "operations_note": text(row[21]),
                    "fabric": text(row[28]),
                    "selling_points": text(row[29]),
                    "colors": text(row[30]),
                    "main_image_count": image_counts["main"],
                    "white_image_count": image_counts["white"],
                    "scene_image_count": image_counts["scene"],
                })
    finally:
        workbook.close()
    return records


def load_eu_sources(path: Path, resolver: AliasResolver) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    eu_catalog: dict[str, dict[str, Any]] = {}
    de_pricing: dict[str, dict[str, Any]] = {}
    try:
        for row in workbook.worksheets[2].iter_rows(min_row=2, max_col=9, values_only=True):
            current = identifier(row[0])
            if not current:
                continue
            old = identifier(row[1])
            resolver.add(current, old)
            eu_catalog[current] = {
                "current": current,
                "old": old,
                "category": text(row[2]),
                "season": text(row[3]),
                "product_id": identifier(row[4]),
                "amazon_url": text(row[6]),
                "weight_g": number(row[7]),
                "cost_cny": number(row[8]),
            }
        for row in workbook.worksheets[3].iter_rows(min_row=2, max_col=29, values_only=True):
            sku = identifier(row[0])
            if not sku:
                continue
            de_pricing[sku] = {
                "sku": sku,
                "normal_price": number(row[1]),
                "product_id": identifier(row[3]),
                "weight_g": number(row[5]),
                "cost_cny": number(row[6]),
                "discount": number(row[7]),
                "promo_price": number(row[8]),
                "profit_cny": number(row[11]),
                "margin": number(row[12]),
            }
    finally:
        workbook.close()
    return eu_catalog, de_pricing


def load_performance(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    result: dict[str, dict[str, Any]] = {}
    try:
        for row in workbook.worksheets[0].iter_rows(min_row=2, max_col=28, values_only=True):
            product_id = identifier(row[0])
            if product_id:
                result[product_id] = {
                    "gmv": number(row[1]),
                    "units_sold": number(row[2]),
                    "orders": number(row[3]),
                    "video_gmv": number(row[12]),
                    "video_units_sold": number(row[13]),
                    "product_card_gmv": number(row[20]),
                    "product_card_units_sold": number(row[21]),
                    "performance_period": "unspecified",
                }
    finally:
        workbook.close()
    return result


def load_inventory(path: Path, market: str, resolver: AliasResolver) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    records: list[dict[str, Any]] = []
    try:
        for row in workbook.worksheets[2].iter_rows(min_row=4, max_col=14, values_only=True):
            source_sku = identifier(row[0])
            if not source_sku:
                continue
            master_sku, conflict = resolver.resolve(source_sku)
            available = number(row[9]) or 0.0
            in_transit = number(row[10]) or 0.0
            records.append({
                "market": market,
                "sku": master_sku,
                "source_sku": source_sku,
                "match_basis": "direct" if source_sku == master_sku else "old_sku_alias",
                "alias_conflict": conflict,
                "snapshot_date": snapshot_date(path),
                "stock_status": text(row[1]),
                "on_hand_stock": number(row[8]) or 0.0,
                "available_stock": available,
                "in_transit_stock": in_transit,
                "available_plus_in_transit": available + in_transit,
                "daily_sales_7d": number(row[11]) or 0.0,
                "color_count": number(row[2]) or 0.0,
                "full_stock_color_count": number(row[3]) or 0.0,
                "sku_count": number(row[12]) or 0.0,
                "zero_available_sku_count": number(row[13]) or 0.0,
                "is_available": available > 0,
                "source_file": str(path),
            })
    finally:
        workbook.close()
    return records


def choose(records: Iterable[dict[str, Any]], field: str) -> Any:
    for record in records:
        value = record.get(field)
        if value not in (None, "", [], set()):
            return value
    return ""


def extract_asin(url: str) -> str:
    match = re.search(r"/dp/([A-Z0-9]{10})", identifier(url))
    return match.group(1) if match else ""


def source_names(paths: Iterable[Path]) -> str:
    return ";".join(sorted({path.name for path in paths}))


def build_catalog(args: argparse.Namespace) -> dict[str, Any]:
    source_paths = {
        "asin": Path(args.asin_map),
        "performance": Path(args.performance),
        "tiktok": Path(args.tiktok_styles),
        "eu": Path(args.eu_pricing),
        "de_inventory": Path(args.de_inventory),
        "us_inventory": Path(args.us_inventory),
    }
    for path in source_paths.values():
        if not path.is_file():
            raise FileNotFoundError(path)

    resolver = AliasResolver()
    styles = load_tiktok_styles(source_paths["tiktok"], resolver)
    eu_catalog, de_pricing = load_eu_sources(source_paths["eu"], resolver)
    performance = load_performance(source_paths["performance"])
    de_inventory = load_inventory(source_paths["de_inventory"], "DE", resolver)
    us_inventory = load_inventory(source_paths["us_inventory"], "US", resolver)
    inventory_rows = sorted([*us_inventory, *de_inventory], key=lambda row: (row["market"], row["sku"], row["source_sku"]))

    style_by_master: dict[str, list[dict[str, Any]]] = defaultdict(list)
    members: dict[str, set[str]] = defaultdict(set)
    for record in styles:
        master, _ = resolver.resolve(record["current"])
        style_by_master[master].append(record)
        members[master].update({master, record["current"]})
        if record["old"]:
            members[master].add(record["old"])

    eu_by_master: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in eu_catalog.values():
        master, _ = resolver.resolve(record["current"])
        eu_by_master[master].append(record)
        members[master].update({master, record["current"]})
        if record["old"]:
            members[master].add(record["old"])

    pricing_by_master: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in de_pricing.values():
        master, _ = resolver.resolve(record["sku"])
        pricing_by_master[master].append(record)
        members[master].update({master, record["sku"]})

    inventory_by_market_master: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in inventory_rows:
        inventory_by_market_master[(record["market"], record["sku"])].append(record)
        members[record["sku"]].update({record["sku"], record["source_sku"]})

    us_masters = {record["sku"] for record in us_inventory} | set(style_by_master)
    de_masters = {record["sku"] for record in de_inventory} | set(eu_by_master) | set(pricing_by_master)
    market_masters = {"US": us_masters, "DE": de_masters}
    tracked_identifiers = {value for master in us_masters | de_masters for value in members.get(master, {master})}
    tracked_identifiers.update(us_masters | de_masters)

    asin_by_master: dict[str, set[str]] = defaultdict(set)
    active_asin_by_master: dict[str, set[str]] = defaultdict(set)
    delisted_asin_by_master: dict[str, set[str]] = defaultdict(set)
    asin_images: dict[str, str] = {}
    asin_owner: dict[str, str] = {}
    asin_conflicts: set[str] = set()
    workbook = load_workbook(source_paths["asin"], read_only=True, data_only=True, keep_links=False)
    try:
        for asin, image_url, is_delisted, item in workbook.worksheets[0].iter_rows(min_row=2, max_col=4, values_only=True):
            asin_id, item_id = identifier(asin), identifier(item)
            if asin_id and item_id:
                previous = asin_owner.setdefault(asin_id, item_id)
                if previous != item_id:
                    asin_conflicts.add(asin_id)
            if not asin_id or item_id not in tracked_identifiers:
                continue
            master, _ = resolver.resolve(item_id)
            asin_by_master[master].add(asin_id)
            if identifier(is_delisted) in {"\u662f", "YES", "Y", "TRUE", "1"}:
                delisted_asin_by_master[master].add(asin_id)
            else:
                active_asin_by_master[master].add(asin_id)
            if text(image_url):
                asin_images.setdefault(asin_id, text(image_url))
    finally:
        workbook.close()

    product_master: list[dict[str, Any]] = []
    identifier_aliases: list[dict[str, Any]] = []
    source_updated_at = max(iso_mtime(path) for path in source_paths.values())

    for market in ("US", "DE"):
        for master in sorted(market_masters[market]):
            style_records = sorted(style_by_master.get(master, []), key=lambda row: row["segment"] != "apparel")
            eu_records = eu_by_master.get(master, [])
            price_records = pricing_by_master.get(master, [])
            stock_records = inventory_by_market_master.get((market, master), [])
            product_ids: set[str] = set()
            product_id_sources: set[str] = set()
            primary_product_id = ""
            if market == "US":
                for record in style_records:
                    product_ids.update(record["product_ids"])
                    product_id_sources.update(record["product_id_sources"])
                    primary_product_id = primary_product_id or record["primary_product_id"]
            else:
                for record in [*price_records, *eu_records]:
                    product_id = identifier(record.get("product_id"))
                    if product_id:
                        product_ids.add(product_id)
                        product_id_sources.add(f"de_product_id:{product_id}")
                primary_product_id = identifier(choose(price_records, "product_id")) or identifier(choose(eu_records, "product_id"))

            selected_performance = performance.get(primary_product_id, {})
            segment = choose(style_records, "segment") or ("eu_apparel" if eu_records else "inventory_only")
            operations_note = text(choose(style_records, "operations_note"))
            category = text(choose(eu_records, "category")) if market == "DE" else ("apparel" if segment == "apparel" else "")
            if segment == "standard_product":
                category = "standard_product"
            normal_price = number(choose(style_records, "normal_price")) if market == "US" else number(choose(price_records, "normal_price"))
            promo_price = number(choose(style_records, "promo_price")) if market == "US" else number(choose(price_records, "promo_price"))
            image_counts = {
                kind: sum(int(record.get(f"{kind}_image_count") or 0) for record in style_records)
                for kind in ("main", "white", "scene")
            }
            available_stock = sum(float(record["available_stock"]) for record in stock_records)
            fields = {
                "title": "",
                "product_type": "",
                "selling_points": text(choose(style_records, "selling_points")),
                "pain_points": "",
                "proof_actions": "",
                "price": promo_price if promo_price is not None else normal_price,
                "image": image_counts["main"] + image_counts["white"] + image_counts["scene"],
            }
            missing = [name for name, value in fields.items() if value in (None, "", 0)]
            usable = bool(fields["price"] and fields["image"] and text(choose(style_records, "fabric")) and fields["selling_points"])
            ready = usable and bool(category or fields["product_type"] or fields["title"]) and bool(fields["proof_actions"])
            catalog_status = "ready_for_semantic_match" if ready else "usable_with_review" if usable else "needs_enrichment"
            source_files = {source_paths["asin"], source_paths["tiktok"]}
            if market == "DE":
                source_files.update({source_paths["eu"], source_paths["de_inventory"], source_paths["performance"]})
            else:
                source_files.add(source_paths["us_inventory"])

            product_master.append({
                "market": market,
                "sku": master,
                "display_name": operations_note or f"SKU {master}",
                "title": "",
                "title_status": "missing",
                "catalog_segment": segment,
                "category": category,
                "product_type": "",
                "planning_type": text(choose(style_records, "planning_type")),
                "style_label": text(choose(style_records, "style_label")),
                "season": text(choose(eu_records, "season")) or text(choose(style_records, "season")),
                "operations_note": operations_note,
                "fabric": text(choose(style_records, "fabric")),
                "colors": text(choose(style_records, "colors")),
                "selling_points": fields["selling_points"],
                "pain_points": "",
                "proof_actions": "",
                "scenarios": "",
                "normal_price": normal_price,
                "promo_price": promo_price,
                "currency": "USD" if market == "US" else "EUR",
                "weight_g": number(choose(price_records, "weight_g")) or number(choose(eu_records, "weight_g")),
                "cost_cny": number(choose(price_records, "cost_cny")) or number(choose(eu_records, "cost_cny")),
                "profit_cny": number(choose(price_records, "profit_cny")),
                "margin": number(choose(price_records, "margin")),
                "gmv": selected_performance.get("gmv"),
                "units_sold": selected_performance.get("units_sold"),
                "video_gmv": selected_performance.get("video_gmv"),
                "product_card_gmv": selected_performance.get("product_card_gmv"),
                "performance_period": selected_performance.get("performance_period", ""),
                "image_workbook": str(source_paths["tiktok"]) if style_records else "",
                "image_sheet": text(choose(style_records, "source_sheet")),
                "image_row": choose(style_records, "source_row"),
                "main_image_count": image_counts["main"],
                "white_image_count": image_counts["white"],
                "scene_image_count": image_counts["scene"],
                "has_market_inventory": bool(stock_records),
                "is_in_stock": available_stock > 0,
                "catalog_status": catalog_status,
                "missing_fields": ";".join(missing),
                "source_updated_at": source_updated_at,
                "source_files": source_names(source_files),
            })

            aliases = set(members.get(master, {master}))
            aliases.discard(master)
            asins = sorted(asin_by_master.get(master, set()))
            active_asins = sorted(active_asin_by_master.get(master, set()))
            delisted_asins = sorted(delisted_asin_by_master.get(master, set()))
            preferred_asin = ""
            if market == "DE":
                preferred_asin = extract_asin(text(choose(eu_records, "amazon_url")))
            preferred_asin = preferred_asin or (active_asins[0] if active_asins else asins[0] if asins else "")
            identifier_aliases.append({
                "market": market,
                "sku": master,
                "old_skus": ";".join(sorted(aliases)),
                "tiktok_product_ids": ";".join(sorted(product_ids)),
                "primary_tiktok_product_id": primary_product_id,
                "product_id_sources": ";".join(sorted(product_id_sources)),
                "amazon_asins": ";".join(asins),
                "preferred_asin": preferred_asin,
                "preferred_asin_image_url": asin_images.get(preferred_asin, ""),
                "active_asin_count": len(active_asins),
                "delisted_asin_count": len(delisted_asins),
                "asin_conflict_count": sum(asin in asin_conflicts for asin in asins),
                "identifier_conflict": any(len(resolver.targets.get(alias, set())) > 1 for alias in aliases),
                "source_files": source_names({source_paths["asin"], source_paths["tiktok"], source_paths["eu"]}),
            })

    product_master.sort(key=lambda row: (row["market"], row["sku"]))
    identifier_aliases.sort(key=lambda row: (row["market"], row["sku"]))
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "product_rows": len(product_master),
        "alias_rows": len(identifier_aliases),
        "inventory_rows": len(inventory_rows),
        "products_by_market": dict(Counter(row["market"] for row in product_master)),
        "in_stock_by_market": dict(Counter(row["market"] for row in product_master if row["is_in_stock"])),
        "catalog_status": {
            market: dict(Counter(row["catalog_status"] for row in product_master if row["market"] == market))
            for market in ("US", "DE")
        },
        "source_files": {name: str(path) for name, path in source_paths.items()},
    }
    return {
        "schema_version": "2.0",
        "summary": summary,
        "tables": {
            "product_master": product_master,
            "identifier_aliases": identifier_aliases,
            "market_inventory": inventory_rows,
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--asin-map", required=True)
    parser.add_argument("--performance", required=True)
    parser.add_argument("--tiktok-styles", required=True)
    parser.add_argument("--eu-pricing", required=True)
    parser.add_argument("--de-inventory", required=True)
    parser.add_argument("--us-inventory", required=True)
    parser.add_argument("--output-json", required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    payload = build_catalog(args)
    output = Path(args.output_json)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
